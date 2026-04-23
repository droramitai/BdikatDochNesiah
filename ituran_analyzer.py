#!/usr/bin/env python3
"""
Ituran Report Analyzer - Phase 1
Classifies vehicle stops as פריקת מכולות, הסעות עובדים, חניה/שבת, or חריג.

Usage:
    python ituran_analyzer.py <excel_file> [threshold_minutes]

    threshold_minutes: minutes at a stop that qualifies as "פריקת מכולות" (default 120)

Work hours are configurable (defaults: 05:00-20:00).
Any activity outside those hours is flagged as "חריג" for manual review.
"""

import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta, time as dtime
from collections import defaultdict


# ─── defaults ────────────────────────────────────────────────────────────────

DEFAULT_WORK_START   = 5    # 05:00
DEFAULT_WORK_END     = 20   # 20:00
DEFAULT_THRESHOLD    = 120  # minutes

# A stop that crosses midnight and lasts > this → overnight parking
MIDNIGHT_PARKING_MIN = timedelta(hours=5)
# A stop lasting more than this on the same calendar day → parking regardless
MAX_WORK_STOP = timedelta(hours=12)

TYPE_UNLOAD    = "פריקת מכולות"
TYPE_TRANSPORT = "הסעות עובדים"
TYPE_PARKING   = "חניה"
TYPE_ANOMALY   = "חריג - מחוץ לשעות עבודה"


# ─── helpers ─────────────────────────────────────────────────────────────────

def td_to_hours(td: timedelta) -> float:
    return round(td.total_seconds() / 3600, 2)


def parse_driver_name(raw) -> str:
    if not raw:
        return ""
    clean = str(raw).replace("נהג קבוע:", "").replace("\n", " ").strip()
    parts = clean.split()
    if parts and parts[-1].isdigit():
        parts = parts[:-1]
    return " ".join(parts)


def within_work_hours(dt: datetime, work_start: int, work_end: int) -> bool:
    return work_start <= dt.hour < work_end


# ─── parsing ─────────────────────────────────────────────────────────────────

def parse_dt(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), "%d/%m/%Y %H:%M:%S")
        except ValueError:
            return None
    return None


def parse_events(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    events = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == "זמן הודעה":
                header_found = True
            continue

        dt = parse_dt(row[0])
        if dt is None:
            continue

        status = str(row[5]).strip(".").strip() if row[5] else ""
        if "סגירת סוויץ" not in status and "פתיחת סוויץ" not in status:
            continue

        events.append({
            "dt":      dt,
            "driver":  parse_driver_name(row[3]),
            "address": str(row[4]).strip() if row[4] else "",
            "km":      float(row[2]) if row[2] else 0.0,
            "status":  status,
        })

    events.sort(key=lambda e: e["dt"])
    return events


# ─── classification ──────────────────────────────────────────────────────────

def classify_stop(start: datetime, end: datetime, duration: timedelta,
                  threshold: timedelta, work_start: int, work_end: int) -> str:
    """
    Determine the type of a stop period.

    Rules (in order):
      1. Very long or clearly overnight → parking
      2. Started outside work hours → anomaly
      3. Normal duration-based classification
    """
    crosses_midnight = start.date() != end.date()

    if duration >= MAX_WORK_STOP:
        return TYPE_PARKING
    if crosses_midnight and duration >= MIDNIGHT_PARKING_MIN:
        return TYPE_PARKING

    if not within_work_hours(start, work_start, work_end):
        return TYPE_ANOMALY

    return TYPE_UNLOAD if duration >= threshold else TYPE_TRANSPORT


def classify_drive(start: datetime, end: datetime,
                   work_start: int, work_end: int) -> str:
    """A drive is anomalous if it starts outside work hours."""
    if not within_work_hours(start, work_start, work_end):
        return TYPE_ANOMALY
    return TYPE_TRANSPORT


def anomaly_reason(item: dict, work_start: int, work_end: int) -> str:
    """Human-readable explanation for why this item is anomalous."""
    h = item["start"].hour
    if h < work_start:
        return (f"נסיעה/עצירה בשעות הלילה ({item['start'].strftime('%H:%M')}) "
                f"- לפני שעת התחלה {work_start:02d}:00")
    return (f"נסיעה/עצירה בשעות הערב ({item['start'].strftime('%H:%M')}) "
            f"- אחרי שעת סיום {work_end:02d}:00")


# ─── build periods ───────────────────────────────────────────────────────────

def build_periods(events: list[dict], threshold: timedelta,
                  work_start: int, work_end: int) -> tuple[list, list]:
    """
    Returns (stops, drives).
    Each stop : {start, end, driver, address, duration, type, date}
    Each drive: {start, end, driver, duration, type, date}
    """
    stops  = []
    drives = []

    i = 0
    while i < len(events):
        ev = events[i]

        if "סגירת סוויץ" in ev["status"]:
            if i + 1 < len(events) and "פתיחת סוויץ" in events[i + 1]["status"]:
                nxt      = events[i + 1]
                duration = nxt["dt"] - ev["dt"]
                driver   = ev["driver"] or nxt["driver"]

                stops.append({
                    "start":    ev["dt"],
                    "end":      nxt["dt"],
                    "driver":   driver,
                    "address":  ev["address"],
                    "duration": duration,
                    "type":     classify_stop(ev["dt"], nxt["dt"], duration,
                                              threshold, work_start, work_end),
                    "date":     ev["dt"].date(),
                })

                if i + 2 < len(events) and "סגירת סוויץ" in events[i + 2]["status"]:
                    after     = events[i + 2]
                    drive_dur = after["dt"] - nxt["dt"]
                    drives.append({
                        "start":        nxt["dt"],
                        "end":          after["dt"],
                        "driver":       nxt["driver"] or after["driver"],
                        "duration":     drive_dur,
                        "type":         classify_drive(nxt["dt"], after["dt"],
                                                       work_start, work_end),
                        "date":         nxt["dt"].date(),
                        "from_address": ev["address"],    # where the vehicle left from
                        "to_address":   after["address"], # where it arrived
                    })

                i += 2
                continue

        i += 1

    return stops, drives


# ─── aggregation ─────────────────────────────────────────────────────────────

def aggregate(stops: list[dict], drives: list[dict]) -> dict:
    """Returns {(driver, date): {unload, transport, drives, parking, anomaly}}"""
    result = defaultdict(lambda: {
        "unload":    timedelta(),
        "transport": timedelta(),
        "drives":    timedelta(),
        "parking":   timedelta(),
        "anomaly":   timedelta(),
    })

    for s in stops:
        key = (s["driver"], s["date"])
        if s["type"] == TYPE_UNLOAD:
            result[key]["unload"] += s["duration"]
        elif s["type"] == TYPE_PARKING:
            result[key]["parking"] += s["duration"]
        elif s["type"] == TYPE_ANOMALY:
            result[key]["anomaly"] += s["duration"]
        else:  # TYPE_TRANSPORT
            result[key]["transport"] += s["duration"]

    for d in drives:
        key = (d["driver"], d["date"])
        if d["type"] == TYPE_ANOMALY:
            result[key]["anomaly"] += d["duration"]
        else:
            result[key]["drives"] += d["duration"]

    return result


# ─── Excel styling ───────────────────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
BODY_FONT = Font(name="Arial", size=10)
CENTER    = Alignment(horizontal="center", vertical="center")
STRIPE    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")

FILL_UNLOAD    = PatternFill("solid", start_color="DCE8F5", end_color="DCE8F5")
FILL_TRANSPORT = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
FILL_PARKING   = PatternFill("solid", start_color="FAFADC", end_color="FAFADC")
FILL_ANOMALY   = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
FILL_WEEKEND   = PatternFill("solid", start_color="F0E8FF", end_color="F0E8FF")
FILL_VACATION  = PatternFill("solid", start_color="E8FFF0", end_color="E8FFF0")

TYPE_FILL = {
    TYPE_UNLOAD:    FILL_UNLOAD,
    TYPE_TRANSPORT: FILL_TRANSPORT,
    TYPE_PARKING:   FILL_PARKING,
    TYPE_ANOMALY:   FILL_ANOMALY,
}

# Display labels matching the screen
DISPLAY_LABELS = {
    ("stop",  TYPE_UNLOAD):    "עבודה",
    ("stop",  TYPE_TRANSPORT): "עצירת ביניים",
    ("stop",  TYPE_PARKING):   "חניה",
    ("stop",  TYPE_ANOMALY):   "חריג",
    ("drive", TYPE_TRANSPORT): "נסיעה",
    ("drive", TYPE_ANOMALY):   "נסיעה",
}
SPECIAL_FILL = {
    "סוף שבוע/חג": FILL_WEEKEND,
    "חופשה":       FILL_VACATION,
}


def style_header(cell, text, width=None, col_letter=None, ws=None):
    cell.value = text
    cell.font  = HDR_FONT
    cell.fill  = HDR_FILL
    cell.alignment = CENTER
    if width and col_letter and ws:
        ws.column_dimensions[col_letter].width = width


def style_body(cell, fill=None):
    cell.font      = BODY_FONT
    cell.alignment = CENTER
    if fill:
        cell.fill = fill


def set_col(ws, col, header, width):
    from openpyxl.utils import get_column_letter
    letter = get_column_letter(col)
    ws.column_dimensions[letter].width = width
    cell = ws.cell(row=1, column=col, value=header)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = CENTER
    return letter


# ─── sheet writers ───────────────────────────────────────────────────────────

HEBREW_DAYS = {0: "שני", 1: "שלישי", 2: "רביעי", 3: "חמישי", 4: "שישי", 5: "שבת", 6: "ראשון"}


def write_summary_sheet(wb, summary: dict, skip_dates: frozenset = frozenset()):
    from openpyxl.utils import get_column_letter
    ws = wb.active
    ws.title = "סיכום"
    ws.sheet_view.rightToLeft = True
    ws.row_dimensions[1].height = 22

    # RTL order: col A appears RIGHTMOST on screen
    cols = [
        ("שם עובד",                    18),   # A – rightmost
        ("תאריך",                       14),   # B
        ("יום",                         10),   # C
        ("שעות עבודה",                  18),   # D
        ("שעות עצירות ביניים",          20),   # E
        ("שעות נסיעה",                  18),   # F
        ('סה"כ',                        14),   # G
        ("שעות חניה (לא נספר)",         20),   # H
        ("שעות חריגות (לבירור)",        20),   # I – leftmost
    ]
    num_cols = len(cols)
    for col, (header, width) in enumerate(cols, 1):
        set_col(ws, col, header, width)

    sorted_keys = sorted(
        (k for k in summary.keys() if k[1] not in skip_dates),
        key=lambda k: (k[1], k[0])
    )
    last_data_row = 1
    for r, key in enumerate(sorted_keys, 2):
        driver, dt = key
        d = summary[key]
        unload    = td_to_hours(d["unload"])
        transport = td_to_hours(d["transport"])
        drv       = td_to_hours(d["drives"])
        parking   = td_to_hours(d["parking"])
        anomaly   = td_to_hours(d["anomaly"])
        total     = round(unload + transport + drv, 2)
        day_name  = HEBREW_DAYS[dt.weekday()]

        row_data = [driver, dt.strftime("%d/%m/%Y"), day_name,
                    unload, transport, drv, total, parking, anomaly]
        fill = STRIPE if r % 2 == 0 else None

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            if col == 9 and anomaly > 0:   # col I = anomaly
                style_body(cell, FILL_ANOMALY)
            else:
                style_body(cell, fill)
        last_data_row = r

    # ── autofilter ──
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{last_data_row}"

    # ── SUBTOTAL totals row ──
    if last_data_row >= 2:
        total_row = last_data_row + 1
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill      = HDR_FILL
            cell.alignment = CENTER
        ws.cell(row=total_row, column=1, value='סה"כ')
        # SUBTOTAL(9,…) respects filters; cols D-I (4-9) are numeric
        for col in range(4, num_cols + 1):
            letter = get_column_letter(col)
            ws.cell(row=total_row, column=col,
                    value=f"=SUBTOTAL(9,{letter}2:{letter}{last_data_row})"
                    ).number_format = "0.00"


def write_detail_sheet(wb, stops: list[dict], drives: list[dict],
                        special_labels: dict = {}):
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("פירוט עצירות")
    ws.sheet_view.rightToLeft = True
    ws.row_dimensions[1].height = 22

    # RTL order: col A = rightmost on screen
    cols = [
        ("שם עובד",              15),   # A
        ("תאריך",                13),   # B
        ("יום",                  10),   # C
        ("שעת הגעה",             12),   # D = start
        ("שעת יציאה",            12),   # E = end
        ("משך (דקות)",           14),   # F
        ("סיווג",                22),   # G
        ("כתובת / מוצא ← יעד",  50),   # H
    ]
    num_cols = len(cols)
    for col, (h, w) in enumerate(cols, 1):
        set_col(ws, col, h, w)

    all_items = (
        [("stop", s) for s in stops] +
        [("drive", d) for d in drives]
    )
    all_items.sort(key=lambda x: x[1]["start"])

    r = 2
    for kind, item in all_items:
        mins     = int(item["duration"].total_seconds() / 60)
        day_name = HEBREW_DAYS[item["date"].weekday()]

        sp_lbl = special_labels.get((item["date"], item["start"]))
        if sp_lbl:
            label = sp_lbl
            fill  = SPECIAL_FILL.get(sp_lbl, None)
        else:
            label = DISPLAY_LABELS.get((kind, item["type"]), item["type"])
            fill  = TYPE_FILL.get(item["type"], None)

        if kind == "drive":
            frm  = item.get("from_address", "") or ""
            to   = item.get("to_address",   "") or ""
            addr = f"{frm} ← {to}" if frm or to else "(נסיעה)"
        else:
            addr = item["address"]

        row_data = [
            item["driver"],
            item["date"].strftime("%d/%m/%Y"),
            day_name,
            item["start"].strftime("%H:%M"),
            item["end"].strftime("%H:%M"),
            mins,
            label,
            addr,
        ]
        for col, val in enumerate(row_data, 1):
            style_body(ws.cell(row=r, column=col, value=val), fill)
        r += 1

    # ── autofilter ──
    if r > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{r - 1}"

    # ── SUBTOTAL totals row ──
    if r > 2:
        last_data_row = r - 1
        total_row = r
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=total_row, column=col, value="")
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill      = HDR_FILL
            cell.alignment = CENTER
        ws.cell(row=total_row, column=1, value='סה"כ')
        # SUBTOTAL(9,…) for minutes (col F = 6)
        ws.cell(row=total_row, column=6,
                value=f"=SUBTOTAL(9,F2:F{last_data_row})")


def write_anomaly_sheet(wb, stops: list[dict], drives: list[dict],
                        work_start: int, work_end: int):
    from openpyxl.utils import get_column_letter
    anomalies = [s for s in stops  if s["type"] == TYPE_ANOMALY] + \
                [d for d in drives if d["type"] == TYPE_ANOMALY]
    if not anomalies:
        return

    ws = wb.create_sheet("חריגים - לבירור")
    ws.sheet_view.rightToLeft = True
    ws.row_dimensions[1].height = 22

    # RTL order: col A = rightmost
    cols = [
        ("שם עובד",      15),   # A
        ("תאריך",        13),   # B
        ("יום",          10),   # C
        ("שעת התחלה",    13),   # D
        ("שעת סיום",     13),   # E
        ("משך (דקות)",   14),   # F
        ("כתובת / סוג",  36),   # G
        ("הערה",         42),   # H
    ]
    num_cols = len(cols)
    for col, (h, w) in enumerate(cols, 1):
        set_col(ws, col, h, w)

    last_r = 1
    for r, item in enumerate(sorted(anomalies, key=lambda x: x["start"]), 2):
        mins     = int(item["duration"].total_seconds() / 60)
        day_name = HEBREW_DAYS[item["date"].weekday()]
        if "from_address" in item:
            frm  = item.get("from_address", "") or ""
            to   = item.get("to_address",   "") or ""
            addr = f"{frm} ← {to}" if frm or to else "(נסיעה)"
        else:
            addr = item.get("address", "")
        note = anomaly_reason(item, work_start, work_end)
        row_data = [
            item["driver"],
            item["date"].strftime("%d/%m/%Y"),
            day_name,
            item["start"].strftime("%H:%M"),
            item["end"].strftime("%H:%M"),
            mins,
            addr,
            note,
        ]
        for col, val in enumerate(row_data, 1):
            style_body(ws.cell(row=r, column=col, value=val), FILL_ANOMALY)
        last_r = r

    # ── autofilter ──
    if last_r > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{last_r}"


def write_params_sheet(wb, filepath: str, threshold_minutes: int,
                       work_start: int, work_end: int,
                       stops: list, drives: list):
    ws = wb.create_sheet("פרמטרים")
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32

    counts = {t: sum(1 for s in stops if s["type"] == t) for t in
              [TYPE_UNLOAD, TYPE_TRANSPORT, TYPE_PARKING, TYPE_ANOMALY]}

    info = [
        ("קובץ מקור",                   os.path.basename(filepath) if isinstance(filepath, str) else filepath),
        ("סף פריקת מכולות",             f"{threshold_minutes} דקות"),
        ("שעת תחילת עבודה",             f"{work_start:02d}:00"),
        ("שעת סיום עבודה",              f"{work_end:02d}:00"),
        ("סף חניה לילה",                f">5 שעות שעוברות חצות"),
        ("סף חניה מקסימלית",            f">12 שעות"),
        ("",                            ""),
        ("סה״כ עצירות",                 len(stops)),
        ("  פריקת מכולות",              counts[TYPE_UNLOAD]),
        ("  הסעות עובדים",              counts[TYPE_TRANSPORT]),
        ("  חניה / שבת",                counts[TYPE_PARKING]),
        ("  חריגים לבירור",             counts[TYPE_ANOMALY]),
        ("סה״כ נסיעות",                 len(drives)),
        ("נסיעות חריגות",               sum(1 for d in drives if d["type"] == TYPE_ANOMALY)),
    ]

    for r, (k, v) in enumerate(info, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=bool(k), name="Arial")
        ws.cell(row=r, column=2, value=v).font = Font(name="Arial")


# ─── main entry ──────────────────────────────────────────────────────────────

def _parse_and_classify(filepath_or_buffer, threshold_minutes, work_start, work_end):
    """Parse file and classify events. Returns (stops, drives, summary)."""
    threshold = timedelta(minutes=threshold_minutes)
    events    = parse_events(filepath_or_buffer)
    stops, drives = build_periods(events, threshold, work_start, work_end)
    summary   = aggregate(stops, drives)
    return stops, drives, summary


def build_excel_buffer(stops, drives, summary, filename,
                       threshold_minutes, work_start, work_end,
                       skip_dates: frozenset = frozenset(),
                       special_labels: dict = {}):
    """Build Excel workbook from pre-parsed data. Returns BytesIO."""
    import io
    wb = openpyxl.Workbook()
    write_summary_sheet(wb, summary, skip_dates)
    write_detail_sheet(wb, stops, drives, special_labels)
    write_anomaly_sheet(wb, stops, drives, work_start, work_end)
    write_params_sheet(wb, filename, threshold_minutes, work_start, work_end, stops, drives)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def analyze_to_buffer(file_buffer, filename: str = "report.xlsx",
                      threshold_minutes: int = DEFAULT_THRESHOLD,
                      work_start: int = DEFAULT_WORK_START,
                      work_end:   int = DEFAULT_WORK_END,
                      skip_dates: frozenset = frozenset()):
    """
    Analyse an in-memory file buffer (e.g. from Streamlit file_uploader).
    Returns (BytesIO, summary_dict, stops, drives).
    """
    stops, drives, summary = _parse_and_classify(
        file_buffer, threshold_minutes, work_start, work_end
    )
    buf = build_excel_buffer(
        stops, drives, summary, filename,
        threshold_minutes, work_start, work_end, skip_dates
    )
    return buf, summary, stops, drives


def analyze(filepath: str,
            threshold_minutes: int = DEFAULT_THRESHOLD,
            work_start: int = DEFAULT_WORK_START,
            work_end:   int = DEFAULT_WORK_END) -> str:
    """CLI entry: read from disk, write result to disk."""
    print(f"Reading: {os.path.basename(filepath)}")
    wb, summary, stops, drives = _build_workbook(
        filepath, os.path.basename(filepath), threshold_minutes, work_start, work_end
    )
    anomaly_stops  = [s for s in stops  if s["type"] == TYPE_ANOMALY]
    anomaly_drives = [d for d in drives if d["type"] == TYPE_ANOMALY]
    print(f"Events parsed | Stops: {len(stops)}  Drives: {len(drives)}")
    if anomaly_stops or anomaly_drives:
        print(f"[!] {len(anomaly_stops)} anomalous stops + {len(anomaly_drives)} anomalous drives")

    base, _ = os.path.splitext(filepath)
    output_path = f"{base}_ניתוח.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python ituran_analyzer.py <excel_file> [threshold_minutes]")
        print("Example: python ituran_analyzer.py report.xlsx 120")
        sys.exit(1)

    filepath = sys.argv[1]
    threshold_minutes = int(sys.argv[2]) if len(sys.argv) > 2 else 120

    if not os.path.exists(filepath):
        print(f"Error: file not found: {filepath}")
        sys.exit(1)

    analyze(filepath, threshold_minutes)


if __name__ == "__main__":
    main()
